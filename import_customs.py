#!/usr/bin/env python3
"""
报关单PDF → Excel 自动化导入工具 v3

用法:
  python3 import_customs.py 报关单.pdf
  python3 import_customs.py 报关单.pdf 模板.xlsx
  python3 import_customs.py 报关单.pdf 模板.xlsx 输出.xlsx
"""
import sys, os, re
from typing import Optional, List, Dict, Tuple, Any
from copy import copy

try:
    import fitz
except ImportError:
    print("❌ 缺少依赖: pip install PyMuPDF")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("❌ 缺少依赖: pip install openpyxl")
    sys.exit(1)

# ============================================================
# 1. PDF 文本提取
# ============================================================
def extract_text_items(pdf_path: str) -> List[Dict]:
    doc = fitz.open(pdf_path)
    items = []
    for page in doc:
        blocks = page.get_text("dict")["blocks"]
        for b in blocks:
            if "lines" not in b:
                continue
            for line in b["lines"]:
                text = "".join([s["text"] for s in line["spans"]]).strip()
                if text:
                    bbox = line["bbox"]
                    items.append({
                        "x": bbox[0], "y": bbox[1],
                        "w": bbox[2] - bbox[0], "h": bbox[3] - bbox[1],
                        "text": text,
                    })
    doc.close()
    items.sort(key=lambda t: (t["y"], t["x"]))
    return items

# ============================================================
# 2. 报关单数据提取 (坐标定位 + 规则匹配)
# ============================================================
def extract_customs_data(pdf_path: str) -> Dict:
    items = extract_text_items(pdf_path)

    result: Dict[str, Any] = {
        "source_file": os.path.basename(pdf_path),
        "commodity": {},
    }

    def find_items_near(xr: Tuple[float,float], yr: Tuple[float,float]) -> List[Dict]:
        found = []
        for it in items:
            if xr[0] <= it["x"] <= xr[1] and yr[0] <= it["y"] <= yr[1]:
                found.append(it)
        found.sort(key=lambda t: t["y"])
        return found

    def first_in(xr, yr, pattern=None) -> Optional[str]:
        for it in find_items_near(xr, yr):
            if pattern is None or re.search(pattern, it["text"]):
                return it["text"]
        return None

    def find_value_by_label(label_text: str, y_below: Tuple[float,float] = (5, 15)) -> Optional[str]:
        """找到标签后, 在其正下方取值"""
        label_item = None
        for it in items:
            if label_text in it["text"]:
                label_item = it
                break
        if not label_item:
            return None
        ly = label_item["y"]
        for it in items:
            if it is label_item:
                continue
            dy = it["y"] - ly
            dx = abs(it["x"] - label_item["x"])
            if y_below[0] <= dy <= y_below[1] and dx < 100:
                t = it["text"]
                if t and len(t) > 0:
                    return t
        return None

    def find_code_by_label(label_text: str, x_range: Tuple[float,float]) -> Optional[str]:
        """找到标签后, 在 x_range 范围内找 (数字) 格式的代码"""
        label_item = None
        for it in items:
            if label_text in it["text"]:
                label_item = it
                break
        if not label_item:
            return None
        ly = label_item["y"]
        for it in items:
            if it is label_item:
                continue
            if ly - 5 < it["y"] < ly + 8 and x_range[0] < it["x"] < x_range[1]:
                m = re.match(r'^\((\d+)\)$', it["text"])
                if m:
                    return m.group(1)
        return None

    # ---- 海关编号 (18位数字) ----
    for it in items:
        if re.match(r'^\d{18}$', it["text"]):
            result["海关编号"] = it["text"]
            break

    # ---- 提运单号 (4字母+数字, y≈124, x≈520) ----
    for it in items:
        if re.match(r'^[A-Z]{4}\d{10,}', it["text"]) and it["y"] > 110:
            result["提运单号"] = it["text"]
            break

    # ---- 境内发货人 (y≈90-110, x<200) ----
    for it in items:
        if 95 < it["y"] < 110 and it["x"] < 200:
            if "有限" in it["text"] or "公司" in it["text"]:
                result["境内发货人"] = it["text"]
                break

    # ---- 海关编码 (企业10位+2字母, 如33029668FY) ----
    for it in items:
        m = re.search(r'\((\d{8,10}[A-Z]{2})\)', it["text"])
        if m:
            result["海关编码"] = m.group(1)
            break

    # ---- 出境关别 + 代码 ----
    out_val = find_value_by_label("出境关别")
    if out_val:
        out_code = find_code_by_label("出境关别", (280, 340))
        result["出境关别"] = f"{out_val} ({out_code})" if out_code else out_val

    # ---- 申报日期 + 出口日期 (y≈85-130, x≈480-620, 8位数字) ----
    for it in items:
        if 85 < it["y"] < 130 and 480 < it["x"] < 620:
            if re.match(r'^\d{8}$', it["text"]):
                dt = it["text"]
                result["申报日期"] = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}"
                break
    # 出口日期默认同申报日期
    if result.get("申报日期") and not result.get("out_date"):
        result["out_date"] = result["申报日期"]

    # ---- 境外收货人 (y≈120-135, x<80) ----
    for it in items:
        if 120 < it["y"] < 135 and it["x"] < 80:
            t = it["text"].strip()
            if t and "有限" not in t and "公司" not in t:
                if not re.match(r'^\(', t):
                    result["境外收货人"] = t
                    break

    # ---- 运输方式 + 代码 ----
    trans_val = find_value_by_label("运输方式")
    if trans_val:
        trans_code = find_code_by_label("运输方式", (280, 320))
        result["运输方式"] = f"{trans_val} ({trans_code})" if trans_code else trans_val

    # ---- 运输工具名称及航次号 (y≈120-130, x≈380-410) ----
    for it in items:
        if 120 < it["y"] < 130 and 380 < it["x"] < 410:
            result["运输工具名称及航次号"] = it["text"]
            break

    # ---- 监管方式 (y≈140-155, x≈250-270) ----
    for it in items:
        if 140 < it["y"] < 155 and 250 < it["x"] < 270:
            result["监管方式"] = it["text"]
            break

    # ---- 合同协议号 (y≈165-180, x<200, 字母数字) ----
    for it in items:
        if 165 < it["y"] < 180 and it["x"] < 200:
            t = it["text"]
            if re.match(r'^[A-Z0-9]{8,}$', t):
                result["合同协议号"] = t
                break

    # ---- 贸易国/运抵国/指运港/离境口岸 (y≈165-175) ----
    for field, xr in [("贸易国", (250,320)), ("运抵国", (380,450)),
                       ("指运港", (510,560)), ("离境口岸", (630,680))]:
        for it in items:
            if 165 < it["y"] < 178 and xr[0] < it["x"] < xr[1]:
                result[field] = it["text"]
                break
    # 离境口岸补充代码
    for it in items:
        if 165 < it["y"] < 178 and 625 < it["x"] < 680:
            t = it["text"]
            if re.match(r'^\(\d+\)$', t):
                if "离境口岸" in result:
                    result["离境口岸"] = f"{result['离境口岸']} {t}"

    # ---- 包装种类 + 代码 ----
    pack_val = find_value_by_label("包装种类")
    if pack_val:
        pack_code = find_code_by_label("包装种类", (55, 100))
        result["包装种类"] = f"{pack_val} ({pack_code})" if pack_code else pack_val

    # ---- 件数/毛重/净重 (y≈188-200) ----
    for field, xr, cast in [("件数", (250,280), int),
                              ("毛重", (295,315), float),
                              ("净重", (375,395), float)]:
        for it in items:
            if 188 < it["y"] < 200 and xr[0] < it["x"] < xr[1]:
                try:
                    result[field] = cast(it["text"])
                except (ValueError, TypeError):
                    result[field] = it["text"] if it["text"] else None
                break

    # ---- 成交方式 (y≈188-200, x≈460-480) ----
    for it in items:
        if 188 < it["y"] < 200 and 460 < it["x"] < 480:
            result["成交方式"] = it["text"]
            break

    # ---- 运费 (y≈188-200, x≈510-600) ----
    for it in items:
        if 188 < it["y"] < 200 and 510 < it["x"] < 600:
            t = it["text"]
            if "USD" in t or re.search(r'/\d+', t):
                result["运费"] = t
                num_m = re.search(r'(\d+(?:\.\d+)?)', t)
                if num_m:
                    result["运费数值"] = float(num_m.group(1))
                break

    # ---- 申报单位 (y≈530-545) ----
    for it in items:
        if 530 < it["y"] < 545 and it["x"] < 300:
            if "报关" in it["text"] or "有限" in it["text"]:
                result["申报单位"] = it["text"]
                break

    # ---- 特殊关系确认 (y≈480-495) ----
    for it in items:
        if 480 < it["y"] < 495 and 620 < it["x"] < 650:
            result["特殊关系确认"] = it["text"]
            break

    # ========================================
    # ---- 商品明细 (y≈280-320) ----
    # ========================================
    commodity = {
        "项号": "", "商品编号": "", "商品名称": "", "规格型号": "",
        "数量": "", "单位": "", "净重": "",
        "单价": "", "总价": "", "币制": "",
        "原产国": "", "最终目的国": "", "境内货源地": "", "征免": "",
    }

    # 商品表列区间 (精确值, 来自PDF坐标)
    COL_RANGES = [
        ("item_no",     0,   55),   # 项号
        ("hs_code",    55,  110),   # 商品编号
        ("name_spec", 110,  343),   # 商品名称及规格型号
        ("qty_unit",  343,  429),   # 数量及单位
        ("price_tot", 429,  510),   # 单价/总价/币制
        ("origin",    510,  575),   # 原产国(地区)
        ("dest",      575,  682),   # 最终目的国(地区)
        ("domestic",  682,  755),   # 境内货源地
        ("tax",       755,  810),   # 征免
    ]

    col_data = {k: [] for k, _, _ in COL_RANGES}
    for it in items:
        if 280 < it["y"] < 320:
            x = it["x"]
            t = it["text"]
            for key, xmin, xmax in COL_RANGES:
                if xmin <= x < xmax:
                    col_data[key].append(t)
                    break

    # -- 项号 --
    for t in col_data["item_no"]:
        m = re.match(r'^(\d+)$', t)
        if m:
            commodity["项号"] = m.group(1)
            break

    # -- 商品编号 + 名称 --
    for t in col_data["hs_code"]:
        m = re.match(r'^(\d{8,12})\s*(.*)', t)
        if m:
            commodity["商品编号"] = m.group(1)
            if m.group(2):
                commodity["商品名称"] = m.group(2)
            break
    if not commodity["商品编号"]:
        for t in col_data["hs_code"]:
            if re.match(r'^\d+$', t):
                commodity["商品编号"] = t
                break

    # -- 规格型号 (含|分隔符的长文本) --
    for t in col_data["name_spec"]:
        if '|' in t and len(t) > 15:
            commodity["规格型号"] = t
            break

    # -- 数量及单位 (多行: 60件 / 936千克 / 60个) --
    qty_parts = []
    for t in col_data["qty_unit"]:
        for part in t.split():
            m = re.match(r'^(\d+(?:\.\d+)?)(.+)$', part)
            if m:
                qty_parts.append((m.group(1), m.group(2)))
    for num, unit in qty_parts:
        if unit in ("个", "件", "台", "套", "辆"):
            commodity["数量"] = num
            commodity["单位"] = unit
        elif unit in ("千克", "kg", "KG"):
            commodity["净重"] = num

    # -- 单价/总价/币制 (多行: 33.7500 / 2025.00 / 美元) --
    for t in col_data["price_tot"]:
        if "美元" in t or "USD" in t.upper():
            commodity["币制"] = "美元"
        else:
            try:
                val = float(t)
                # 单价通常有较多小数位(如33.7500), 总价通常为整数或两位小数
                if "." in t and len(t.split(".")[-1]) >= 3:
                    commodity["单价"] = val
                elif not commodity.get("单价"):
                    commodity["单价"] = val
                else:
                    commodity["总价"] = val
            except ValueError:
                pass

    # -- 原产国 (中国 + CHN) --
    origin_parts = []
    for t in col_data["origin"]:
        if "中国" in t: origin_parts.insert(0, "中国")
        if "CHN" in t.upper(): origin_parts.append("(CHN)")
    if origin_parts:
        commodity["原产国"] = "".join(origin_parts)

    # -- 最终目的国 (美国 + USA) --
    dest_parts = []
    for t in col_data["dest"]:
        if "美国" in t: dest_parts.insert(0, "美国")
        if "USA" in t.upper(): dest_parts.append("(USA)")
    if dest_parts:
        commodity["最终目的国"] = "".join(dest_parts)

    # -- 境内货源地 --
    for t in col_data["domestic"]:
        if "宁波" in t or "其他" in t:
            commodity["境内货源地"] = t
            break

    # -- 征免 --
    tax_parts = []
    for t in col_data["tax"]:
        if "照章" in t or "征税" in t or "免税" in t or "保证金" in t:
            tax_parts.insert(0, t)
        elif re.match(r'^\(\d+\)$', t):
            tax_parts.append(t)
    if tax_parts:
        commodity["征免"] = " ".join(tax_parts)

    # -- 类型转换 --
    for key in ["数量", "单价", "总价", "净重"]:
        v = commodity.get(key)
        if v and isinstance(v, str):
            try:
                commodity[key] = float(v) if "." in v else int(v)
            except ValueError:
                pass

    result["commodity"] = commodity
    return result

# ============================================================
# 3. 写入 Excel 模板
# ============================================================
COLUMN_MAP = [
    # (列号, 数据key, 备注)
    (1,  "source_file",        "文件名"),
    (2,  "境内发货人",          None),
    (3,  "海关编码",            None),
    (4,  "出境关别",            None),
    (5,  "out_date",           "出口日期"),
    (6,  "申报日期",            None),
    (7,  "境外收货人",          None),
    (8,  "运输方式",            None),
    (9,  "运输工具名称及航次号", None),
    (10, "提运单号",            None),
    (11, "监管方式",            None),
    (12, "合同协议号",          None),
    (13, "离境口岸",            None),
    (14, "包装种类",            None),
    (15, "件数",               None),
    (16, "毛重",               None),
    (17, "净重",               None),
    (18, "成交方式",            None),
    (19, "运费数值",            None),
    (20, "申报单位",            None),
    (21, "commodity.项号",      None),
    (22, "commodity.商品编号",   None),
    (23, "commodity_spec",      "商品名称及规格型号"),
    (24, "commodity.数量",      None),
    (25, "commodity.单位",      None),
    (26, "commodity.单价",      None),
    (27, "commodity.总价",      None),
    (28, "commodity.原产国",     None),
    (29, "commodity.境内货源地", None),
    (30, "commodity.最终目的国", None),
    (31, "commodity.币制",      None),
]

def format_commodity_spec(comm: Dict) -> str:
    """格式化商品名称及规格型号"""
    spec = comm.get("规格型号", "")
    if not spec:
        name = comm.get("商品名称", "")
        return name if name else ""
    # 规格格式: 0|2|放东西用|铁管+密度板|无品牌|93×47×75cm
    parts = spec.split("|")
    labels = ["品牌类型", "出口享惠", "用途", "材质", "品牌", "规格尺寸"]
    desc = comm.get("商品名称", "")
    if desc:
        desc += "\n"
    for i, p in enumerate(parts):
        if i < len(labels):
            desc += f"{labels[i]}:{p}\n"
        else:
            desc += f"{p}\n"
    return desc.strip()

def fill_template(data: Dict, template_path: str, output_path: str):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Sheet1"]

    ref_cell = ws.cell(2, 1)
    data_font = copy(ref_cell.font) if ref_cell.font else Font(name='宋体', size=10)
    data_align = copy(ref_cell.alignment) if ref_cell.alignment else Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    data_fill = copy(ref_cell.fill) if ref_cell.fill else PatternFill()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # 找最后有数据的行
    last_row = 2
    max_col = max(col for col, _, _ in COLUMN_MAP)
    for row in range(2, ws.max_row + 1):
        if any(ws.cell(row, c).value is not None for c in range(1, max_col + 1)):
            last_row = row
    write_row = last_row + 1

    commodity = data.get("commodity", {})

    col_values = {}
    for col, key, _desc in COLUMN_MAP:
        value = None
        if key.startswith("commodity."):
            sub_key = key[len("commodity."):]
            value = commodity.get(sub_key)
        elif key == "commodity_spec":
            value = format_commodity_spec(commodity)
        else:
            value = data.get(key)
        if value is not None and value != "":
            col_values[col] = value

    for col, value in col_values.items():
        cell = ws.cell(write_row, col, value=value)
        cell.font = data_font
        cell.alignment = data_align
        cell.fill = data_fill
        cell.border = thin_border

    ws.row_dimensions[write_row].height = 38
    wb.save(output_path)
    return write_row

# ============================================================
# 4. 主入口
# ============================================================
def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    pdf_path = sys.argv[1]
    template_path = sys.argv[2] if len(sys.argv) > 2 else None
    output_path = sys.argv[3] if len(sys.argv) > 3 else None

    if template_path is None:
        # 在脚本同目录下查找模板文件
        script_dir = os.path.dirname(os.path.abspath(__file__))
        candidates = [f for f in os.listdir(script_dir) if f.endswith('.xlsx')]
        if candidates:
            template_path = os.path.join(script_dir, candidates[0])
        else:
            print("❌ 未找到模板文件，请指定模板路径: python3 import_customs.py 报关单.pdf 模板.xlsx")
            sys.exit(1)

    if output_path is None:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        output_dir = os.path.dirname(pdf_path) or os.path.dirname(template_path)
        output_path = os.path.join(output_dir, f"{base}_导入结果.xlsx")

    if not os.path.exists(pdf_path):
        print(f"❌ PDF 不存在: {pdf_path}")
        sys.exit(1)
    if not os.path.exists(template_path):
        print(f"❌ 模板不存在: {template_path}")
        sys.exit(1)

    print(f"📄 读取 PDF: {pdf_path}")
    data = extract_customs_data(pdf_path)

    print(f"\n📋 提取结果:")
    for k, v in data.items():
        if k not in ("commodity",):
            print(f"  {k}: {v}")

    comm = data.get("commodity", {})
    if comm:
        print(f"\n📦 商品明细:")
        for k, v in comm.items():
            if v:
                print(f"  {k}: {v}")

    print(f"\n📝 写入模板: {template_path}")
    row = fill_template(data, template_path, output_path)
    print(f"✅ 完成! 写入第 {row} 行")
    print(f"📁 输出: {output_path}")

if __name__ == "__main__":
    main()
